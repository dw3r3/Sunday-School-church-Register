from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from datetime import timedelta, datetime, date
import calendar
from flask_sqlalchemy import SQLAlchemy
import os
import shutil
import zipfile
import glob
from apscheduler.schedulers.background import BackgroundScheduler
from werkzeug.utils import secure_filename
from PIL import Image
from functools import wraps
import pandas as pd
from io import BytesIO

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Accept either a session 'user' or explicit 'logged_in' flag for compatibility
        if not session.get('user') and not session.get('logged_in'):
            flash('You need to be logged in as an admin to access this page.', 'error')
            return redirect(url_for('home'))
        if session.get('role') != 'admin':
            flash('Admin access required.', 'error')
            return redirect(url_for('home'))
        return f(*args, **kwargs)
    return decorated_function

# -------------------------------
# Backup Management
# -------------------------------

# Initialize scheduler
scheduler = BackgroundScheduler()
scheduler.start()

def create_backup():
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_dir = os.path.join(app.root_path, 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        backup_path = os.path.join(backup_dir, f'backup_{timestamp}.zip')
        
        with zipfile.ZipFile(backup_path, 'w') as zipf:
            # Backup database
            db_path = os.path.join(app.root_path, 'instance', 'database.db')
            zipf.write(db_path, 'database.db')
            
            # Backup profile pictures
            profiles_dir = os.path.join(app.root_path, 'static', 'uploads', 'profiles')
            if os.path.exists(profiles_dir):
                for file in os.listdir(profiles_dir):
                    file_path = os.path.join(profiles_dir, file)
                    zipf.write(file_path, os.path.join('profiles', file))
        
        # Update last backup time
        config = BackupConfig.query.first()
        if config:
            config.last_backup = datetime.now()
            db.session.commit()
            
        cleanup_old_backups()
        return True
    except Exception as e:
        print(f"Backup failed: {str(e)}")
        return False

def cleanup_old_backups():
    config = BackupConfig.query.first()
    if not config:
        return
    
    backup_dir = os.path.join(app.root_path, 'backups')
    if not os.path.exists(backup_dir):
        return
        
    backup_files = glob.glob(os.path.join(backup_dir, 'backup_*.zip'))
    backup_files.sort(key=os.path.getmtime, reverse=True)
    
    # Keep only the most recent backups based on max_backups setting
    if len(backup_files) > config.max_backups:
        for old_backup in backup_files[config.max_backups:]:
            try:
                os.remove(old_backup)
            except Exception as e:
                print(f"Failed to remove old backup {old_backup}: {str(e)}")

def schedule_backups():
    config = BackupConfig.query.first()
    if not config:
        config = BackupConfig()
        db.session.add(config)
        db.session.commit()
    
    if config.auto_backup_enabled:
        if config.backup_frequency == 'daily':
            scheduler.add_job(
                create_backup,
                'interval',
                days=1,
                id='daily_backup',
                replace_existing=True
            )
        elif config.backup_frequency == 'weekly':
            scheduler.add_job(
                create_backup,
                'interval',
                weeks=1,
                id='weekly_backup',
                replace_existing=True
            )
    else:
        # Remove any existing backup jobs
        scheduler.remove_all_jobs()

# -------------------------------
# Flask App Config
# -------------------------------
app = Flask(__name__)
app.secret_key = "your_secret_key"
app.permanent_session_lifetime = timedelta(minutes=30)

# -------------------------------
# MySQL DB Config
# -------------------------------
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///church_register.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Image upload configuration
UPLOAD_FOLDER = 'static/uploads/profiles'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Create upload directory if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def resize_image(image_path, max_size=(200, 200)):
    """Resize image to max_size while maintaining aspect ratio"""
    try:
        with Image.open(image_path) as img:
            img.thumbnail(max_size, Image.Resampling.LANCZOS)
            img.save(image_path, optimize=True, quality=85)
    except Exception as e:
        print(f"Error resizing image: {e}")

db = SQLAlchemy(app)

# -------------------------------
# Student Model (Table)
# -------------------------------
class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    dob = db.Column(db.String(20), nullable=False)
    parent = db.Column(db.String(100))
    contact = db.Column(db.String(50))
    student_class = db.Column(db.String(50))
    status = db.Column(db.String(10), default="active")
    deletion_requested = db.Column(db.Boolean, default=False)
    profile_image = db.Column(db.String(200))  # Store image filename
    family_id = db.Column(db.String(50))  # For grouping family members

         #Attebndance Model
class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    present = db.Column(db.Boolean, default=True)

    # Relationship (optional but powerful)
    student = db.relationship('Student', backref=db.backref('attendances', lazy=True))

# -------------------------------
# Inventory Model (Table)
# -------------------------------
class Inventory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_name = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Integer, default=0)
    description = db.Column(db.String(200))
    date_added = db.Column(db.DateTime, default=datetime.now)
    last_checked = db.Column(db.DateTime, default=datetime.now)
    notes = db.Column(db.Text)  # For missing items explanations

class InventoryAudit(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('inventory.id'), nullable=False)
    action = db.Column(db.String(50), nullable=False)  # 'added', 'deleted', 'found', 'missing'
    date = db.Column(db.DateTime, default=datetime.now)
    user = db.Column(db.String(100))
    notes = db.Column(db.Text)

    # Relationship
    item = db.relationship('Inventory', backref=db.backref('audit_logs', lazy=True))

# -------------------------------
# Backup Configuration
# -------------------------------
class BackupConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    auto_backup_enabled = db.Column(db.Boolean, default=False)
    backup_frequency = db.Column(db.String(20), default='weekly')  # daily, weekly
    last_backup = db.Column(db.DateTime, nullable=True)
    max_backups = db.Column(db.Integer, default=10)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

# -------------------------------
# User Model
# -------------------------------
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="teacher")
    full_name = db.Column(db.String(100), nullable=True)
    email = db.Column(db.String(100), unique=True, nullable=True)
    phone = db.Column(db.String(20), nullable=True)
    assigned_class = db.Column(db.String(50), nullable=True)
    preferred_class = db.Column(db.String(50), nullable=True)
    status = db.Column(db.String(20), nullable=False, default="pending")
    registration_message = db.Column(db.Text, nullable=True)
    approved_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    approved_at = db.Column(db.DateTime, nullable=True)
    last_login = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.now)

# -------------------------------
# Initialize Default Users
# -------------------------------
def create_default_users():
    with app.app_context():
        # Check if admin exists
        admin = User.query.filter_by(username='admin@church.org').first()
        if not admin:
            admin = User(
                username='admin@church.org',
                password='admin123',
                role='admin',
                full_name='System Administrator',
                email='admin@church.org',
                status='active',
                assigned_class=None
            )
            db.session.add(admin)

        # Check if teacher exists
        teacher = User.query.filter_by(username='teacher@church.org').first()
        if not teacher:
            teacher = User(
                username='teacher@church.org',
                password='teacher123',
                role='teacher',
                full_name='Demo Teacher',
                email='teacher@church.org',
                status='active',
                assigned_class='Genesis'
            )
            db.session.add(teacher)

        db.session.commit()

# -------------------------------
# Helper: Get all Sundays in a month
# -------------------------------
def get_sundays(year, month):
    sundays = []
    cal = calendar.Calendar()
    for day in cal.itermonthdates(year, month):
        if day.weekday() == 6 and day.month == month:
            sundays.append(day)
    return sundays

# -------------------------------
# Jinja Filter + Now Context
# -------------------------------
@app.template_filter('todatetime')
def to_datetime_filter(s):
    return datetime.strptime(s, "%Y-%m-%d")

@app.context_processor
def inject_now():
    return {'now': datetime.now}

# -------------------------------
# Login Page
# -------------------------------
@app.route("/")
def home():
    return render_template("login.html")

# -------------------------------
# Login Submission
# -------------------------------
@app.route("/login", methods=["POST"])
def login():
    email = request.form["email"]
    password = request.form["password"]

    # Check database users first
    user = User.query.filter_by(email=email, password=password).first()
    if not user:
        # Fallback to old users dict for existing accounts
        user = User.query.filter_by(username=email, password=password).first()

    if user:
        if user.status == 'pending':
            flash("Your account is pending admin approval. Please wait for confirmation.", "warning")
            return redirect(url_for("home"))
        elif user.status == 'rejected':
            flash("Your account has been rejected. Please contact the administrator.", "error")
            return redirect(url_for("home"))
        elif user.status == 'suspended':
            flash("Your account has been suspended. Please contact the administrator.", "error")
            return redirect(url_for("home"))
        elif user.status == 'active':
            # Update last login
            user.last_login = datetime.now()
            db.session.commit()

            session["user"] = email
            session["user_id"] = user.id
            session["role"] = user.role
            session["assigned_class"] = user.assigned_class
            session["full_name"] = user.full_name
            flash("Login successful!", "success")
            return redirect(url_for("dashboard"))

    # Fallback to old users dict for backward compatibility
    if email in users and users[email]["password"] == password:
        session["user"] = email
        session["role"] = users[email]["role"]
        return redirect(url_for("dashboard"))

    flash("Invalid login credentials!", "error")
    return redirect(url_for("home"))

# -------------------------------
# Teacher Registration
# -------------------------------
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        # Get form data
        full_name = request.form.get("full_name", "").strip()
        email = request.form.get("email", "").strip()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        preferred_class = request.form.get("preferred_class", "")
        phone = request.form.get("phone", "").strip()
        message = request.form.get("message", "").strip()

        # Validation
        if not all([full_name, email, username, password, preferred_class]):
            flash("Please fill in all required fields.", "error")
            return render_template("register.html")

        if password != confirm_password:
            flash("Passwords do not match.", "error")
            return render_template("register.html")

        if len(password) < 6:
            flash("Password must be at least 6 characters long.", "error")
            return render_template("register.html")

        # Check for existing users
        if User.query.filter_by(email=email).first():
            flash("Email address already registered.", "error")
            return render_template("register.html")

        if User.query.filter_by(username=username).first():
            flash("Username already taken.", "error")
            return render_template("register.html")

        # Create new user
        new_user = User(
            full_name=full_name,
            email=email,
            username=username,
            password=password,  # In production, hash this password
            role="teacher",
            preferred_class=preferred_class,
            phone=phone,
            registration_message=message,
            status="pending"
        )

        try:
            db.session.add(new_user)
            db.session.commit()

            flash("Registration successful! Your account is pending admin approval. You will be notified once approved.", "success")
            return redirect(url_for("home"))

        except Exception as e:
            db.session.rollback()
            flash("Registration failed. Please try again.", "error")
            return render_template("register.html")

    # GET request - show registration form
    classes = ["Genesis", "Exodus", "Psalms", "Proverbs", "Revelation", "High Schoolers"]
    return render_template("register.html", classes=classes)

# -------------------------------
# Dashboard
# -------------------------------
@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():
    if "user" not in session:
        return redirect(url_for("home"))

    user_role = session.get("role")
    assigned_class = session.get("assigned_class")

    today = date.today()
    month = int(request.args.get("month", today.month))
    year = int(request.args.get("year", today.year))
    selected_class = request.args.get("class_name")

    # For teachers, allow viewing all classes but default to their assigned class
    if user_role == "teacher" and assigned_class and not selected_class:
        selected_class = assigned_class

    sundays = get_sundays(year, month)

    # Find the current Sunday (today if it's Sunday, or the most recent Sunday)
    current_sunday = None
    if today.weekday() == 6:  # Today is Sunday
        current_sunday = today
    else:
        # Find the most recent Sunday in the current month's Sundays
        for sunday in reversed(sundays):
            if sunday <= today:
                current_sunday = sunday
                break

    # If no current Sunday found in this month, use today's date
    if current_sunday is None:
        current_sunday = today

    # Support searching by family id (family number) in addition to class
    family_id = request.args.get("family_id")

    query = Student.query.filter_by(status="active")
    if selected_class:
        query = query.filter_by(student_class=selected_class)
    if family_id:
        # family_id stored as string in the model; filter using provided value
        query = query.filter_by(family_id=family_id)

    filtered_students = query.all()

    # Check for students at risk of deactivation (for admin notification)
    at_risk_count = 0
    if session.get("role") == "admin":
        # Get the last 4 Sundays for at-risk check
        today = datetime.now()
        check_sundays = []
        current_date = today

        for _ in range(4):
            days_back = (current_date.weekday() + 1) % 7
            if days_back == 0 and current_date.date() == today.date():
                days_back = 7

            sunday = current_date - timedelta(days=days_back)
            check_sundays.append(sunday.date())
            current_date = sunday - timedelta(days=1)

        # Count students at risk (all active students, not just filtered)
        all_active_students = Student.query.filter_by(status='active').all()
        for student in all_active_students:
            missed_count = 0
            for sunday in check_sundays:
                attendance = Attendance.query.filter_by(student_id=student.id, date=sunday).first()
                if not attendance or not attendance.present:
                    missed_count += 1

            if missed_count >= 3:  # At risk if missed 3+ Sundays
                at_risk_count += 1

    return render_template("dashboard.html",
                           students=filtered_students,
                           sundays=sundays,
                           month=month,
                           year=year,
                           selected_class=selected_class,
                           current_sunday=current_sunday,
                           at_risk_count=at_risk_count,
                           family_id=family_id)

# -------------------------------
# Add Student
# -------------------------------
@app.route("/add_student", methods=["POST"])
def add_student():
    name = request.form["name"]
    dob = request.form.get("dob", "").strip()
    age_input = request.form.get("age", "").strip()
    parent = request.form.get("parent", "")
    contact = request.form.get("contact", "")
    family_id = request.form.get("family_id", "")

    today = datetime.now()

    # Determine birth date from either dob or age input. dob takes precedence.
    birth = None
    if dob:
        try:
            birth = datetime.strptime(dob, "%Y-%m-%d")
        except Exception:
            flash("Invalid date of birth format. Use YYYY-MM-DD.", "error")
            return redirect(url_for("dashboard"))
        if birth > today:
            flash("Date of birth cannot be in the future!", "error")
            return redirect(url_for("dashboard"))
        age = (today - birth).days // 365
    elif age_input:
        try:
            age_int = int(age_input)
            if age_int < 0 or age_int > 120:
                raise ValueError
        except Exception:
            flash("Invalid age provided.", "error")
            return redirect(url_for("dashboard"))
        # Approximate birth date: set to today's date minus age years
        try:
            birth = today.replace(year=today.year - age_int)
        except Exception:
            # Fallback if replace fails (e.g., Feb 29)
            birth = today - timedelta(days=age_int * 365)
        age = age_int
        # Store dob as YYYY-MM-DD string for the Student model
        dob = birth.strftime("%Y-%m-%d")
    else:
        flash("Please provide either a date of birth or an age.", "error")
        return redirect(url_for("dashboard"))

    if age <= 5:
        assigned_class = "Genesis"
    elif age <= 7:
        assigned_class = "Exodus"
    elif age <= 9:
        assigned_class = "Psalms"
    elif age <= 11:
        assigned_class = "Proverbs"
    elif age <= 13:
        assigned_class = "Revelation"
    else:
        assigned_class = "High Schoolers"

    # Handle profile image upload
    profile_image_filename = None
    if 'profile_image' in request.files:
        file = request.files['profile_image']
        if file and file.filename != '' and allowed_file(file.filename):
            # Create unique filename
            timestamp = int(datetime.now().timestamp())
            filename = secure_filename(f"{timestamp}_{file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

            try:
                file.save(filepath)
                # Resize image to save space
                resize_image(filepath)
                profile_image_filename = filename
            except Exception as e:
                flash(f"Error uploading image: {str(e)}", "warning")

    student = Student(
        name=name,
        dob=dob,
        parent=parent,
        contact=contact,
        student_class=assigned_class,
        profile_image=profile_image_filename,
        family_id=family_id if family_id else None
    )

    db.session.add(student)
    db.session.commit()

    flash("Student registered successfully!", "success")
    return redirect(url_for("dashboard"))



# -------------------------------
# Logout
# -------------------------------
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))

# -------------------------------
# Request Deletion (Teacher)
# -------------------------------
@app.route("/delete_request/<int:student_id>", methods=["POST"])
def mark_for_deletion(student_id):
    if session.get("role") != "teacher":
        flash("Only teachers can request deletion", "error")
        return redirect(url_for("dashboard"))

    student = Student.query.get(student_id)
    if student and student.status == "active":
        student.deletion_requested = True
        db.session.commit()
        flash("Deletion request sent to admin", "success")

    return redirect(url_for("dashboard"))

# -------------------------------
# Approve Delete (Admin)
# -------------------------------
@app.route("/approve_delete/<int:student_id>", methods=["POST"])
def approve_delete(student_id):
    if session.get("role") != "admin":
        flash("Only admins can approve deletion", "error")
        return redirect(url_for("dashboard"))

    student = Student.query.get(student_id)
    if student:
        student.status = "deleted"
        student.deletion_requested = False
        db.session.commit()
        flash("Student marked as deleted", "success")

    return redirect(url_for("dashboard"))

# -------------------------------
# Reject Delete (Admin)
# -------------------------------
@app.route("/reject_delete/<int:student_id>", methods=["POST"])
def reject_delete(student_id):
    if session.get("role") != "admin":
        flash("Only admins can reject deletion", "error")
        return redirect(url_for("dashboard"))

    student = Student.query.get(student_id)
    if student:
        student.deletion_requested = False
        db.session.commit()
        flash("Deletion request rejected", "success")

    return redirect(url_for("dashboard"))

   #ttendance Marking

@app.route("/mark_attendance", methods=["POST"])
def mark_attendance():
    if "user" not in session:
        return redirect(url_for("home"))

    student_id = request.form.get("student_id")
    date_str = request.form.get("date")
    present = request.form.get("present") == "true"

    date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()

    # Check if already marked
    record = Attendance.query.filter_by(student_id=student_id, date=date_obj).first()
    if record:
        record.present = present
    else:
        record = Attendance(student_id=student_id, date=date_obj, present=present)
        db.session.add(record)

    db.session.commit()
    return "Attendance marked", 200


@app.context_processor
def utility_functions():
    def attendance_present(student_id, sunday):
        record = Attendance.query.filter_by(student_id=student_id, date=sunday).first()
        return record.present if record else False
    return dict(attendance_present=attendance_present)


@app.route("/attendance_report")
def attendance_report():
    if "user" not in session:
        return redirect(url_for("home"))

    today = date.today()
    month = int(request.args.get("month", today.month))
    year = int(request.args.get("year", today.year))
    selected_class = request.args.get("class_name")

    sundays = get_sundays(year, month)

    # Fetch students
    if selected_class:
        students = Student.query.filter_by(student_class=selected_class, status="active").all()
    else:
        students = Student.query.filter_by(status="active").all()

    # Fetch all attendance records for this month
    all_attendance = Attendance.query.filter(
        Attendance.date.between(date(year, month, 1), date(year, month, 31))
    ).all()

    # Create quick lookup: {(student_id, date): present}
    attendance_map = {
        (a.student_id, a.date): a.present for a in all_attendance
    }

    return render_template("attendance_report.html", students=students,
                           sundays=sundays, month=month, year=year,
                           selected_class=selected_class, attendance_map=attendance_map)


@app.route('/get_student/<int:student_id>')
def get_student(student_id):
    if "user" not in session:
        return {"error": "Unauthorized"}, 401

    student = Student.query.get_or_404(student_id)
    return {
        "id": student.id,
        "name": student.name,
        "dob": student.dob,
        "parent": student.parent,
        "contact": student.contact,
        "student_class": student.student_class,
        "family_id": student.family_id,
        "profile_image": student.profile_image
    }

@app.route('/edit_student', methods=['GET', 'POST'])
def edit_student():
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    student_id = request.form.get("student_id")
    student = Student.query.get_or_404(student_id)

    # Update basic info
    student.name = request.form.get("name")
    student.dob = request.form.get("dob")
    student.parent = request.form.get("parent", "")
    student.contact = request.form.get("contact", "")
    student.family_id = request.form.get("family_id", "") or None

    # Handle profile image update
    if 'profile_image' in request.files:
        file = request.files['profile_image']
        if file and file.filename != '' and allowed_file(file.filename):
            # Delete old image if exists
            if student.profile_image:
                old_path = os.path.join(app.config['UPLOAD_FOLDER'], student.profile_image)
                if os.path.exists(old_path):
                    os.remove(old_path)

            # Save new image
            timestamp = int(datetime.now().timestamp())
            filename = secure_filename(f"{timestamp}_{file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

            try:
                file.save(filepath)
                resize_image(filepath)
                student.profile_image = filename
            except Exception as e:
                flash(f"Error uploading image: {str(e)}", "warning")

    # Recalculate class based on age
    birth = datetime.strptime(student.dob, "%Y-%m-%d")
    today = datetime.now()
    age = (today - birth).days // 365

    if age <= 5:
        student.student_class = "Genesis"
    elif age <= 7:
        student.student_class = "Exodus"
    elif age <= 9:
        student.student_class = "Psalms"
    elif age <= 11:
        student.student_class = "Proverbs"
    elif age <= 13:
        student.student_class = "Revelation"
    else:
        student.student_class = "High Schoolers"

    db.session.commit()
    flash(f"Student '{student.name}' updated successfully!", "success")
    return redirect(url_for("dashboard"))

@app.route("/student/<int:student_id>")
def student_detail(student_id):
    if "user" not in session:
        flash("Please log in first.", "error")
        return redirect(url_for("home"))

    student = Student.query.get_or_404(student_id)

    # Calculate age
    from datetime import datetime
    birth_date = datetime.strptime(student.dob, "%Y-%m-%d")
    today = datetime.now()
    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

    # Format date for display
    formatted_date = birth_date.strftime("%B %d, %Y")

    return render_template("student_detail.html",
                         student=student,
                         age=age,
                         formatted_date=formatted_date)

@app.route("/all_students")
def all_students():
    if "user" not in session:
        flash("You must be logged in to view this page.", "error")
        return redirect(url_for("home"))

    # Get filter parameters
    selected_class = request.args.get("class_name")
    group_by = request.args.get("group_by")
    sort_by = request.args.get("sort_by", "name")
    search = request.args.get("search", "").strip()
    
    class_list = ['Genesis', 'Exodus', 'Psalms', 'Proverbs', 'Revelation', 'High Schoolers']

    # Base query
    query = Student.query.filter_by(status="active")
    
    # Apply class filter if selected
    if selected_class:
        query = query.filter_by(student_class=selected_class)
    
    # Get all students for statistics
    all_students = query.all()
    
    # Calculate statistics
    total_students = len(all_students)
    total_families = len(set(s.family_id for s in all_students if s.family_id))
    selected_class_count = len([s for s in all_students if not selected_class or s.student_class == selected_class])
    
    # Apply search filter if provided
    if search:
        search = search.lower()
        students = [s for s in all_students if
                   search in s.name.lower() or
                   (s.parent and search in s.parent.lower()) or
                   (s.family_id and search in s.family_id.lower())]
    else:
        students = all_students
    
    # Sorting functions
    def name_sort_key(s):
        return s.name.lower() if s.name else ''

    def family_sort_key(s):
        # First sort by whether they have a family ID
        has_family = 0 if s.family_id else 1
        # Then by the family ID (numeric if possible)
        try:
            family_num = int(s.family_id) if s.family_id else float('inf')
            family_key = (has_family, family_num)
        except (ValueError, TypeError):
            family_key = (has_family, s.family_id or '')
        # Finally by name
        return (family_key, name_sort_key(s))

    def class_sort_key(s):
        # Define class order based on typical age progression
        class_order = {
            'Genesis': 1,
            'Exodus': 2,
            'Psalms': 3,
            'Proverbs': 4,
            'Revelation': 5,
            'High Schoolers': 6
        }
        return (class_order.get(s.student_class, 99), name_sort_key(s))

    def dob_sort_key(s):
        try:
            # Parse date string to datetime object for proper comparison
            dob = datetime.strptime(s.dob, "%Y-%m-%d")
            return (dob, name_sort_key(s))
        except (ValueError, TypeError):
            # Handle invalid dates by putting them at the end
            return (datetime.max, name_sort_key(s))

    # Apply sorting based on selected option
    if sort_by == 'family':
        students = sorted(students, key=family_sort_key)
    elif sort_by == 'class':
        students = sorted(students, key=class_sort_key)
    elif sort_by == 'dob':
        students = sorted(students, key=dob_sort_key)
    else:  # sort_by == 'name'
        students = sorted(students, key=name_sort_key)

    # If requested, group students by family_id while maintaining sort order
    grouped_families = None
    if group_by == 'family':
        # First, group students by family while preserving their order
        families = {}
        ordered_families = []
        
        for s in students:
            key = s.family_id if s.family_id else 'No Family'
            if key not in families:
                families[key] = []
                ordered_families.append(key)
            families[key].append(s)

        # For each family group, maintain the sort order of students within
        for family in families.values():
            if sort_by == 'class':
                family.sort(key=class_sort_key)
            elif sort_by == 'dob':
                family.sort(key=dob_sort_key)
            else:
                family.sort(key=name_sort_key)

        # Create the final ordered list of (family_id, students) tuples
        def family_group_sort_key(family_id):
            if family_id == 'No Family':
                return (1, float('inf'))
            try:
                return (0, int(family_id))
            except (ValueError, TypeError):
                return (0, str(family_id))

        # Create final sorted list of family groups
        if sort_by == 'family':
            # When sorting by family, sort the family groups themselves
            ordered_families.sort(key=family_group_sort_key)
        
        grouped_families = [(k, families[k]) for k in ordered_families]

    return render_template("all_students.html",
                         students=students,
                         selected_class=selected_class,
                         class_list=class_list,
                         grouped_families=grouped_families,
                         group_by=group_by,
                         sort_by=sort_by,
                         search=search,
                         total_students=total_students,)

@app.route("/download_students")
def download_students():
    if "user" not in session:
        flash("You must be logged in to download data.", "error")
        return redirect(url_for("home"))

    # Get filter parameters - same as all_students view
    selected_class = request.args.get("class_name")
    group_by = request.args.get("group_by")
    sort_by = request.args.get("sort_by", "name")
    search = request.args.get("search", "").strip()
    
    # Base query
    query = Student.query.filter_by(status="active")
    
    # Apply class filter if selected
    if selected_class:
        query = query.filter_by(student_class=selected_class)
    
    # Get filtered students
    students = query.all()
    
    # Apply search filter if provided
    if search:
        students = [s for s in students if search.lower() in s.name.lower()]

    # Sort students based on sort_by parameter
    if sort_by == "name":
        students.sort(key=lambda x: x.name.lower())
    elif sort_by == "class":
        students.sort(key=lambda x: (x.student_class, x.name.lower()))
    elif sort_by == "age":
        students.sort(key=lambda x: (x.birth_date if x.birth_date else date.max))
    elif sort_by == "family":
        students.sort(key=lambda x: (str(x.family_id), x.name.lower()))

    # Group students by family
    families = {}
    for student in students:
        family_id = student.family_id or 'No Family ID'
        if family_id not in families:
            families[family_id] = {
                'students': [],
                'parent': student.parent,
                'contact': student.contact
            }
        families[family_id]['students'].append(student)

    # Prepare data for Excel with family grouping
    data = []
    for family_id, family in sorted(families.items()):
        # Add family header
        data.append({
            'Family ID': f"Family {family_id}",
            'Name': '',
            'Class': '',
            'Birth Date': '',
            'Parent': family['parent'] or 'N/A',
            'Contact': family['contact'] or 'N/A'
        })
        # Add students in the family
        for student in sorted(family['students'], key=lambda x: x.name.lower()):
            data.append({
                'Family ID': '',  # Empty for students in family
                'Name': f"  • {student.name}",  # Indent student names
                'Class': student.student_class,
                'Birth Date': student.dob,
                'Parent': '',  # Empty for students in family
                'Contact': ''  # Empty for students in family
            })
        # Add a blank row between families
        data.append({
            'Family ID': '',
            'Name': '',
            'Class': '',
            'Birth Date': '',
            'Parent': '',
            'Contact': ''
        })
    df = pd.DataFrame(data)
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Students')
        
        # Get the worksheet
        worksheet = writer.sheets['Students']
        
        # Auto-adjust columns width
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            ) + 2
            worksheet.column_dimensions[chr(65 + idx)].width = max_length
            
        # Format the worksheet
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Define styles
        header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        family_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        bold_font = Font(bold=True)
        
        # Format header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = bold_font
        
        # Format family rows and their students
        row = 2  # Start after header
        while row <= len(data):
            cell = worksheet.cell(row=row, column=1)
            if cell.value and cell.value.startswith('Family'):
                # Family header row
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.fill = family_fill
                    cell.font = bold_font
            row += 1
    
    output.seek(0)
    
    return send_file(
        output,
        download_name=f'students_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/inventory')
def inventory():
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    items = Inventory.query.all()

    # Get all unique categories dynamically
    categories = set()
    for item in items:
        if item.description and " - QR: " in item.description:
            category = item.description.split(" - QR: ")[0]
            # Skip placeholder items when showing categories
            if not item.item_name.endswith(" Placeholder"):
                categories.add(category)

    # Organize items by category
    all_items = items
    items_by_category = {}

    for category in categories:
        items_by_category[category] = [
            item for item in items
            if item.description and item.description.startswith(category + " - QR: ")
        ]

    return render_template("inventory.html",
                         all_items=all_items,
                         categories=sorted(categories),
                         items_by_category=items_by_category)

@app.route('/add_item', methods=['POST'])
def add_item():
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    name = request.form.get("name")
    item_type = request.form.get("type")
    custom_type = request.form.get("custom_type")
    qr_code = request.form.get("qr_code")

    # Use custom type if selected
    if item_type == "Custom" and custom_type:
        item_type = custom_type.strip()

    if name and item_type:
        # If no QR code provided, generate a unique one
        if not qr_code or qr_code.strip() == "":
            import time
            import random
            timestamp = int(time.time())
            random_num = random.randint(100, 999)
            qr_code = f"AUTO_{timestamp}_{random_num}"

        # Check if QR code already exists
        existing_item = Inventory.query.filter(
            Inventory.description.like(f"%QR: {qr_code}%")
        ).first()

        if existing_item:
            flash(f"Item with QR code '{qr_code}' already exists!", "error")
            return redirect(url_for("inventory"))

        # Create new inventory item
        new_item = Inventory(
            item_name=name,
            quantity=1,
            description=f"{item_type} - QR: {qr_code}",
            date_added=datetime.now(),
            last_checked=datetime.now()
        )

        db.session.add(new_item)
        db.session.commit()

        # Log the action
        audit_log = InventoryAudit(
            item_id=new_item.id,
            action='added',
            user=session.get('user', 'Unknown'),
            notes=f"Item added via {'QR scan' if not qr_code.startswith('AUTO_') else 'manual entry'}"
        )
        db.session.add(audit_log)
        db.session.commit()

        if qr_code.startswith("AUTO_"):
            flash(f"Item '{name}' added successfully with auto-generated ID: {qr_code}!", "success")
        else:
            flash(f"Item '{name}' added successfully with QR code: {qr_code}!", "success")
    else:
        flash("Item name and type are required!", "error")

    return redirect(url_for("inventory"))

@app.route('/delete_item/<int:item_id>', methods=['POST'])
def delete_item(item_id):
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    item = Inventory.query.get_or_404(item_id)
    item_name = item.item_name

    try:
        # Log the deletion before deleting
        audit_log = InventoryAudit(
            item_id=item.id,
            action='deleted',
            user=session.get('user', 'Unknown'),
            notes=f"Item '{item_name}' deleted by admin"
        )
        db.session.add(audit_log)

        db.session.delete(item)
        db.session.commit()
        flash(f"Item '{item_name}' deleted successfully!", "success")
    except Exception as e:
        flash(f"Error deleting item: {str(e)}", "error")

    return redirect(url_for("inventory"))

@app.route('/add_category', methods=['POST'])
def add_category():
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    category_name = request.form.get("category_name", "").strip()

    if not category_name:
        flash("Category name is required!", "error")
        return redirect(url_for("inventory"))

    # Create a placeholder item for the new category to make the tab appear
    placeholder_item = Inventory(
        item_name=f"{category_name} Placeholder",
        quantity=0,
        description=f"{category_name} - QR: PLACEHOLDER_{int(datetime.now().timestamp())}",
        date_added=datetime.now(),
        last_checked=datetime.now(),
        notes="Placeholder item to create category tab. Add real items to this category."
    )

    db.session.add(placeholder_item)
    db.session.commit()

    flash(f"Category '{category_name}' added successfully! You can now add items to this category.", "success")
    return redirect(url_for("inventory"))

@app.route('/promote_students', methods=['GET', 'POST'])
def promote_students():
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    if request.method == 'POST':
        promotion_type = request.form.get('promotion_type')

        if promotion_type == 'automatic':
            # Automatic promotion based on age
            students = Student.query.filter_by(status='active').all()
            promoted_count = 0

            for student in students:
                birth_date = datetime.strptime(student.dob, "%Y-%m-%d")
                today = datetime.now()
                age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

                # Determine correct class based on current age
                if age <= 5:
                    new_class = "Genesis"
                elif age <= 7:
                    new_class = "Exodus"
                elif age <= 9:
                    new_class = "Psalms"
                elif age <= 11:
                    new_class = "Proverbs"
                elif age <= 13:
                    new_class = "Revelation"
                else:
                    new_class = "High Schoolers"

                # Only update if class needs to change
                if student.student_class != new_class:
                    old_class = student.student_class
                    student.student_class = new_class
                    promoted_count += 1

                    # Log the promotion
                    print(f"Promoted {student.name} from {old_class} to {new_class} (Age: {age})")

            db.session.commit()
            flash(f"Successfully promoted {promoted_count} students to age-appropriate classes!", "success")

        elif promotion_type == 'manual':
            # Manual promotion for selected students
            selected_students = request.form.getlist('student_ids')
            new_class = request.form.get('new_class')

            if not selected_students or not new_class:
                flash("Please select students and a target class.", "error")
                return redirect(url_for('promote_students'))

            promoted_count = 0
            for student_id in selected_students:
                student = Student.query.get(student_id)
                if student:
                    old_class = student.student_class
                    student.student_class = new_class
                    promoted_count += 1
                    print(f"Manually moved {student.name} from {old_class} to {new_class}")

            db.session.commit()
            flash(f"Successfully moved {promoted_count} students to {new_class}!", "success")

        return redirect(url_for('promote_students'))

    # GET request - show promotion interface
    students = Student.query.filter_by(status='active').all()

    # Calculate age and suggested class for each student
    student_data = []
    for student in students:
        birth_date = datetime.strptime(student.dob, "%Y-%m-%d")
        today = datetime.now()
        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

        # Determine suggested class based on age
        if age <= 5:
            suggested_class = "Genesis"
        elif age <= 7:
            suggested_class = "Exodus"
        elif age <= 9:
            suggested_class = "Psalms"
        elif age <= 11:
            suggested_class = "Proverbs"
        elif age <= 13:
            suggested_class = "Revelation"
        else:
            suggested_class = "High Schoolers"

        needs_promotion = student.student_class != suggested_class

        student_data.append({
            'student': student,
            'age': age,
            'current_class': student.student_class,
            'suggested_class': suggested_class,
            'needs_promotion': needs_promotion
        })

    # Sort by those needing promotion first
    student_data.sort(key=lambda x: (not x['needs_promotion'], x['student'].name))

    classes = ["Genesis", "Exodus", "Psalms", "Proverbs", "Revelation", "High Schoolers"]

    return render_template("promote_students.html",
                         student_data=student_data,
                         classes=classes)

@app.route('/manage_status', methods=['GET', 'POST'])
def manage_status():
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    if request.method == 'POST':
        action = request.form.get('action')
        selected_students = request.form.getlist('student_ids')

        if not selected_students:
            flash("Please select at least one student.", "error")
            return redirect(url_for('manage_status'))

        updated_count = 0
        for student_id in selected_students:
            student = Student.query.get(student_id)
            if student:
                if action == 'activate':
                    student.status = 'active'
                elif action == 'deactivate':
                    student.status = 'inactive'
                updated_count += 1

        db.session.commit()
        status_text = "activated" if action == 'activate' else "deactivated"
        flash(f"Successfully {status_text} {updated_count} students!", "success")

        return redirect(url_for('manage_status'))

    # GET request - show status management interface
    students = Student.query.all()
    active_students = [s for s in students if s.status == 'active']
    inactive_students = [s for s in students if s.status == 'inactive']

    return render_template("manage_status.html",
                         active_students=active_students,
                         inactive_students=inactive_students)

@app.route('/check_attendance_deactivation', methods=['POST'])
def check_attendance_deactivation():
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    # Get the last 4 Sundays
    today = datetime.now()
    sundays = []
    current_date = today

    # Find the last 4 Sundays
    for _ in range(4):
        # Go back to find the most recent Sunday
        days_back = (current_date.weekday() + 1) % 7
        if days_back == 0 and current_date.date() == today.date():
            days_back = 7  # If today is Sunday, go back to previous Sunday

        sunday = current_date - timedelta(days=days_back)
        sundays.append(sunday.date())
        current_date = sunday - timedelta(days=1)  # Move to day before this Sunday

    # Get all active students
    active_students = Student.query.filter_by(status='active').all()
    deactivated_count = 0
    deactivated_students = []

    for student in active_students:
        # Check attendance for the last 4 Sundays
        missed_count = 0
        attendance_details = []

        for sunday in sundays:
            attendance = Attendance.query.filter_by(
                student_id=student.id,
                date=sunday
            ).first()

            if not attendance or not attendance.present:
                missed_count += 1
                attendance_details.append(f"{sunday.strftime('%m/%d/%Y')}: Absent")
            else:
                attendance_details.append(f"{sunday.strftime('%m/%d/%Y')}: Present")

        # If student missed all 4 Sundays, deactivate them
        if missed_count >= 4:
            student.status = 'inactive'
            deactivated_count += 1
            deactivated_students.append({
                'name': student.name,
                'class': student.student_class,
                'attendance': attendance_details
            })

            print(f"Auto-deactivated {student.name} for missing {missed_count}/4 Sundays")

    db.session.commit()

    if deactivated_count > 0:
        flash(f"Automatically deactivated {deactivated_count} students for missing 4+ consecutive Sundays. They can be reactivated when they return.", "warning")

        # Log the deactivations for admin review
        for student_info in deactivated_students:
            print(f"DEACTIVATED: {student_info['name']} ({student_info['class']})")
            for attendance_line in student_info['attendance']:
                print(f"  - {attendance_line}")
    else:
        flash("No students needed automatic deactivation based on attendance.", "info")

    return redirect(url_for("manage_status"))

@app.route('/auto_attendance_check')
def auto_attendance_check():
    """Automatic check that can be called periodically"""
    if not session.get("role") == "admin":
        return {"error": "Access denied"}, 403

    # Get the last 4 Sundays
    today = datetime.now()
    sundays = []
    current_date = today

    for _ in range(4):
        days_back = (current_date.weekday() + 1) % 7
        if days_back == 0 and current_date.date() == today.date():
            days_back = 7

        sunday = current_date - timedelta(days=days_back)
        sundays.append(sunday.date())
        current_date = sunday - timedelta(days=1)

    active_students = Student.query.filter_by(status='active').all()
    students_at_risk = []

    for student in active_students:
        missed_count = 0
        recent_attendance = []

        for sunday in sundays:
            attendance = Attendance.query.filter_by(
                student_id=student.id,
                date=sunday
            ).first()

            if not attendance or not attendance.present:
                missed_count += 1
                recent_attendance.append({"date": sunday.strftime('%m/%d'), "present": False})
            else:
                recent_attendance.append({"date": sunday.strftime('%m/%d'), "present": True})

        if missed_count >= 3:  # At risk if missed 3+ (will be deactivated at 4)
            students_at_risk.append({
                'id': student.id,
                'name': student.name,
                'class': student.student_class,
                'missed_count': missed_count,
                'attendance': recent_attendance,
                'will_deactivate': missed_count >= 4
            })

    return {
        'students_at_risk': students_at_risk,
        'total_at_risk': len(students_at_risk)
    }

# -------------------------------
# Admin Teacher Management
# -------------------------------
@app.route('/admin/teachers', methods=['GET'])
def admin_teachers():
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    # Get all users
    pending_teachers = User.query.filter_by(status='pending', role='teacher').all()
    active_teachers = User.query.filter_by(status='active', role='teacher').all()
    suspended_teachers = User.query.filter_by(status='suspended', role='teacher').all()
    rejected_teachers = User.query.filter_by(status='rejected', role='teacher').all()

    # Get class assignment overview
    classes = ["Genesis", "Exodus", "Psalms", "Proverbs", "Revelation", "High Schoolers"]
    class_assignments = {}
    for class_name in classes:
        teacher = User.query.filter_by(assigned_class=class_name, status='active').first()
        student_count = Student.query.filter_by(student_class=class_name, status='active').count()
        class_assignments[class_name] = {
            'teacher': teacher,
            'student_count': student_count
        }

    return render_template("admin_teachers.html",
                         pending_teachers=pending_teachers,
                         active_teachers=active_teachers,
                         suspended_teachers=suspended_teachers,
                         rejected_teachers=rejected_teachers,
                         class_assignments=class_assignments,
                         classes=classes)

@app.route('/admin/approve-teacher/<int:user_id>', methods=['POST'])
def approve_teacher(user_id):
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    user = User.query.get_or_404(user_id)
    assigned_class = request.form.get('assigned_class')

    if not assigned_class:
        flash("Please select a class to assign the teacher.", "error")
        return redirect(url_for('admin_teachers'))

    # Check if class already has a teacher
    existing_teacher = User.query.filter_by(assigned_class=assigned_class, status='active').first()
    if existing_teacher and existing_teacher.id != user_id:
        flash(f"Class {assigned_class} already has a teacher assigned ({existing_teacher.full_name}). Please reassign or choose a different class.", "error")
        return redirect(url_for('admin_teachers'))

    user.status = 'active'
    user.assigned_class = assigned_class
    user.approved_by = session.get('user_id')
    user.approved_at = datetime.now()

    db.session.commit()

    flash(f"Teacher {user.full_name} approved and assigned to {assigned_class} class!", "success")
    return redirect(url_for('admin_teachers'))

@app.route('/admin/reject-teacher/<int:user_id>', methods=['POST'])
def reject_teacher(user_id):
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    user = User.query.get_or_404(user_id)
    rejection_reason = request.form.get('rejection_reason', '')

    user.status = 'rejected'
    user.approved_by = session.get('user_id')
    user.approved_at = datetime.now()
    if rejection_reason:
        user.registration_message = f"REJECTED: {rejection_reason}"

    db.session.commit()

    flash(f"Teacher registration for {user.full_name} has been rejected.", "warning")
    return redirect(url_for('admin_teachers'))

@app.route('/admin/suspend-teacher/<int:user_id>', methods=['POST'])
def suspend_teacher(user_id):
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    user = User.query.get_or_404(user_id)
    user.status = 'suspended'

    db.session.commit()

    flash(f"Teacher {user.full_name} has been suspended.", "warning")
    return redirect(url_for('admin_teachers'))

@app.route('/admin/reactivate-teacher/<int:user_id>', methods=['POST'])
def reactivate_teacher(user_id):
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    user = User.query.get_or_404(user_id)
    assigned_class = request.form.get('assigned_class')

    if not assigned_class:
        flash("Please select a class to assign the teacher.", "error")
        return redirect(url_for('admin_teachers'))

    user.status = 'active'
    user.assigned_class = assigned_class

    db.session.commit()

    flash(f"Teacher {user.full_name} has been reactivated and assigned to {assigned_class}!", "success")
    return redirect(url_for('admin_teachers'))

@app.route('/admin/reassign-teacher/<int:user_id>', methods=['POST'])
def reassign_teacher(user_id):
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    user = User.query.get_or_404(user_id)
    new_class = request.form.get('new_class')

    if not new_class:
        flash("Please select a new class.", "error")
        return redirect(url_for('admin_teachers'))

    old_class = user.assigned_class
    user.assigned_class = new_class

    db.session.commit()

    flash(f"Teacher {user.full_name} reassigned from {old_class} to {new_class}!", "success")
    return redirect(url_for('admin_teachers'))

# -------------------------------
# Admin Student Deletion
# -------------------------------
@app.route('/admin/delete-student/<int:student_id>', methods=['POST'])
def delete_student(student_id):
    if not session.get("role") == "admin":
        flash("Access denied. Only admins can delete students.", "danger")
        return redirect(url_for("dashboard"))

    student = Student.query.get_or_404(student_id)
    student_name = student.name
    student_class = student.student_class

    try:
        # Delete associated attendance records first (to maintain referential integrity)
        Attendance.query.filter_by(student_id=student_id).delete()

        # Delete the student
        db.session.delete(student)
        db.session.commit()

        flash(f"Student '{student_name}' from {student_class} class has been permanently deleted.", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting student: {str(e)}", "error")

    return redirect(url_for("dashboard", class_name=student_class))

@app.route('/generate_report')
def generate_report():
    if not session.get("role") == "admin":
        flash("Access denied.", "danger")
        return redirect(url_for("dashboard"))

    items = Inventory.query.all()
    total_items = len(items)
    available_items = len([item for item in items if item.quantity > 0])
    missing_items = total_items - available_items

    report_data = {
        'total_items': total_items,
        'available_items': available_items,
        'missing_items': missing_items,
        'items': items
    }

    flash(f"Report generated: {available_items} available, {missing_items} missing out of {total_items} total items.", "info")
    return redirect(url_for("inventory"))

def inventory_item_details(item):
    """Helper function to get item details"""
    qr_code = item.description.split(' - QR: ')[1] if ' - QR: ' in item.description else 'N/A'
    category = item.description.split(' - QR: ')[0] if ' - QR: ' in item.description else item.description
    status = "Available" if item.quantity > 0 else "Missing"
    return qr_code, category, status
# -------------------------------
# Backup and Restore
# -------------------------------
def create_backup():
    """Create a zip file containing the database and profile images"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_dir = os.path.join(app.root_path, 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    
    backup_filename = f'church_register_backup_{timestamp}.zip'
    backup_path = os.path.join(backup_dir, backup_filename)
    
    with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Backup database
        db_path = os.path.join(app.root_path, 'church_register.db')
        if os.path.exists(db_path):
            zipf.write(db_path, 'church_register.db')
        
        # Backup profile images
        profile_dir = os.path.join(app.root_path, 'static/uploads/profiles')
        if os.path.exists(profile_dir):
            for root, _, files in os.walk(profile_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.join('profiles', file)
                    zipf.write(file_path, arcname)
    
    return backup_path

def restore_from_backup(backup_file):
    """Restore database and profile images from a backup zip file"""
    with zipfile.ZipFile(backup_file, 'r') as zipf:
        # Restore database
        if 'church_register.db' in zipf.namelist():
            db_path = os.path.join(app.root_path, 'church_register.db')
            # Optionally create a backup of current DB before overwriting
            if os.path.exists(db_path):
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                shutil.copy2(db_path, f"{db_path}.{timestamp}.bak")
            zipf.extract('church_register.db', app.root_path)
        
        # Restore profile images
        profile_dir = os.path.join(app.root_path, 'static/uploads/profiles')
        os.makedirs(profile_dir, exist_ok=True)
        for file_info in zipf.filelist:
            if file_info.filename.startswith('profiles/'):
                zipf.extract(file_info, os.path.join(app.root_path, 'static/uploads'))

@app.route('/admin/backup', methods=['GET', 'POST'])
def backup_restore():
    if "user" not in session or session.get("role") != "admin":
        flash("Access denied.", "error")
        return redirect(url_for("home"))

    if request.method == 'POST':
        if 'restore_file' in request.files:
            file = request.files['restore_file']
            if file and file.filename.endswith('.zip'):
                try:
                    restore_from_backup(file)
                    flash("Backup restored successfully!", "success")
                except Exception as e:
                    flash(f"Error restoring backup: {str(e)}", "error")
            else:
                flash("Please upload a valid backup file (ZIP)", "error")
        else:
            try:
                backup_path = create_backup()
                return send_file(
                    backup_path,
                    as_attachment=True,
                    download_name=os.path.basename(backup_path)
                )
            except Exception as e:
                flash(f"Error creating backup: {str(e)}", "error")
        
        return redirect(url_for('backup_restore'))

    # List existing backups in the backups directory
    backup_dir = os.path.join(app.root_path, 'backups')
    backups = []
    if os.path.exists(backup_dir):
        backups = sorted(
            [f for f in os.listdir(backup_dir) if f.endswith('.zip')],
            reverse=True
        )

    return render_template(
        'backup_restore.html',
        backups=backups
    )

@app.route('/admin/backup/download/<filename>')
def download_backup(filename):
    if "user" not in session or session.get("role") != "admin":
        flash("Access denied.", "error")
        return redirect(url_for("home"))
        
    backup_dir = os.path.join(app.root_path, 'backups')
    return send_file(
        os.path.join(backup_dir, filename),
        as_attachment=True,
        download_name=filename
    )

# -------------------------------
# Backup Routes
# -------------------------------
@app.route('/backup', methods=['POST'])
@admin_required
def backup():
    try:
        if create_backup():
            flash('Backup created successfully!', 'success')
        else:
            flash('Backup failed!', 'error')
        return redirect(url_for('backup_settings'))
    except Exception as e:
        flash(f'Backup failed: {str(e)}', 'error')
        return redirect(url_for('backup_settings'))

@app.route('/restore', methods=['POST'])
@admin_required
def restore():
    if 'backup_file' not in request.files:
        flash('No backup file selected', 'error')
        return redirect(url_for('backup_settings'))
    
    backup_file = request.files['backup_file']
    if backup_file.filename == '':
        flash('No backup file selected', 'error')
        return redirect(url_for('backup_settings'))
    
    try:
        temp_dir = os.path.join(app.root_path, 'temp_restore')
        os.makedirs(temp_dir, exist_ok=True)
        
        # Extract backup to temp directory
        backup_file.save(os.path.join(temp_dir, 'backup.zip'))
        with zipfile.ZipFile(os.path.join(temp_dir, 'backup.zip'), 'r') as zipf:
            zipf.extractall(temp_dir)
        
        # Restore database
        db_path = os.path.join(app.root_path, 'instance', 'database.db')
        temp_db_path = os.path.join(temp_dir, 'database.db')
        if os.path.exists(temp_db_path):
            shutil.copy2(temp_db_path, db_path)
        
        # Restore profile pictures
        profiles_dir = os.path.join(app.root_path, 'static', 'uploads', 'profiles')
        temp_profiles_dir = os.path.join(temp_dir, 'profiles')
        if os.path.exists(temp_profiles_dir):
            if os.path.exists(profiles_dir):
                shutil.rmtree(profiles_dir)
            shutil.copytree(temp_profiles_dir, profiles_dir)
        
        # Cleanup
        shutil.rmtree(temp_dir)
        
        flash('Backup restored successfully!', 'success')
        return redirect(url_for('backup_settings'))
    except Exception as e:
        flash(f'Restore failed: {str(e)}', 'error')
        return redirect(url_for('backup_settings'))
    finally:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)

@app.route('/backup_settings', methods=['GET', 'POST'])
@admin_required
def backup_settings():
    config = BackupConfig.query.first()
    if not config:
        config = BackupConfig()
        db.session.add(config)
        db.session.commit()
    
    if request.method == 'POST':
        config.auto_backup_enabled = 'auto_backup' in request.form
        config.backup_frequency = request.form.get('backup_frequency', 'weekly')
        config.max_backups = int(request.form.get('max_backups', 10))
        db.session.commit()
        
        schedule_backups()
        flash('Backup settings updated successfully!', 'success')
        return redirect(url_for('backup_settings'))
    
    backup_dir = os.path.join(app.root_path, 'backups')
    backups = []
    if os.path.exists(backup_dir):
        backup_files = glob.glob(os.path.join(backup_dir, 'backup_*.zip'))
        backup_files.sort(key=os.path.getmtime, reverse=True)
        backups = [os.path.basename(f) for f in backup_files]
    
    return render_template('backup_settings.html', config=config, backups=backups)


@app.route('/backup/delete/<filename>', methods=['POST'])
@admin_required
def delete_backup(filename):
    """Delete a backup file from the backups directory (admin only)."""
    backup_dir = os.path.join(app.root_path, 'backups')
    safe_name = secure_filename(filename)
    target_path = os.path.join(backup_dir, safe_name)

    # Prevent path traversal by ensuring target is inside backup_dir
    try:
        if not os.path.abspath(target_path).startswith(os.path.abspath(backup_dir)):
            flash('Invalid backup filename.', 'error')
            return redirect(url_for('backup_restore'))

        if os.path.exists(target_path):
            try:
                os.remove(target_path)
                flash(f'Backup {safe_name} deleted successfully.', 'success')
            except Exception as e:
                flash(f'Failed to delete backup: {str(e)}', 'error')
        else:
            flash('Backup file not found.', 'error')
    except Exception as e:
        flash(f'Error deleting backup: {str(e)}', 'error')

    return redirect(url_for('backup_restore'))

# -------------------------------
# Run App
# -------------------------------
if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        create_default_users()
        
        # Initialize backup config and schedule backups
        if BackupConfig.query.first() is None:
            default_config = BackupConfig()
            db.session.add(default_config)
            db.session.commit()
        schedule_backups()
        
    app.run(debug=True)
